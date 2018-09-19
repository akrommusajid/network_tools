import sys
import textfsm
from pprint import pprint
from openpyxl import load_workbook
from openpyxl import Workbook
from netmiko import ConnectHandler
import re
import time
from datetime import datetime


def time_log():
    time_log_stamp = datetime.now()
    log_time = time_log_stamp.strftime('%Y-%m-%d %H:%M:%S')
    return log_time

def log(write, filename):
    with open(filename, 'a') as writelog:
        output = '%s :' % time_log()
        output += write
        writelog.write(output)

def inventory(filename):
    wb = load_workbook(filename)
    ws = wb.active
    invents = list()
    for i in range(2, ws.max_row+1):
        invent_dic = {
            'hostname': None,
            'address': None,
            'type' : None,
            'user': None,
            'pass': None
        }

        invent_dic['hostname'] = ws.cell(row=i, column=1).value
        invent_dic['address'] = ws.cell(row=i, column=2).value
        invent_dic['type'] = ws.cell(row=i, column=3).value
        invent_dic['user'] = ws.cell(row=i, column=4).value
        invent_dic['pass'] = ws.cell(row=i, column=5).value
        invents.append(invent_dic)
    return invents

def arp_addr(output):
    arp_re = re.compile(r'\d+\.\d+\.\d+\.\d+')
    arp = arp_re.findall(output)
    return arp

class SendCommand(object):
    def __init__(self, hostname, dev_type, address, username, password):
        self.hostname = hostname
        self.dev_type = dev_type
        self.address = address
        self.username = username
        self.password = password

    def command(self, cmd):
        self.cmd = cmd
        connecthandler = {
            'device_type' : self.dev_type,
            'ip' : self.address,
            'username' : self.username,
            'password' : self.password
        }

        netconnect = ConnectHandler(**connecthandler)
        result = netconnect.send_command(self.cmd)
        netconnect.disconnect()
        if len(result) == 0:
            return None
        else:
            return result

    def up_interface(self):
        interface_status_cmd = 'show interface status | exclude Po|disabled|notconnect'
        interface_status_output = self.command(interface_status_cmd)

        interface_re = re.compile(r'Te\S+|Gi\S+|Fa\S+')
        interface = interface_re.findall(interface_status_output)

        return interface

    def traffic_interface(self):
        up_interfaces = self.up_interface()
        interfaces_rate = list()
        interface_input_re = re.compile(r'input rate (\d+)')
        interface_output_re = re.compile(r'output rate (\d+)')

        for interface in up_interfaces:
            
            interface_rate = {
                    'interface' : None,
                    'input' : None,
                    'output' : None
                    }

            interface_rate_cmd = 'show interface %s | i rate' % interface
            interface_rate_output = self.command(interface_rate_cmd)
            interface_input_rate = interface_input_re.search(interface_rate_output)
            interface_output_rate = interface_output_re.search(interface_rate_output)
            if interface_input_rate:
                interface_rate['input'] = interface_input_rate.group(1)
            else:
                pass
            
            if interface_output_rate:
                interface_rate['output'] = interface_output_rate.group(1)
            else:
                pass

            interface_rate['interface'] = interface
            interfaces_rate.append(interface_rate)

        return interfaces_rate

    def mac_port(self):
        trk_list = self.trunk_port()
        if self.dev_type == 'cisco_xe':
            list_trk = list()
            for trk in trk_list:
                te_re = re.compile(r'Te[\d\/]+')
                te = te_re.search(trk)
                ge_re = re.compile(r'Gi[\d\/]+')
                ge = ge_re.search(trk)
                fa_re = re.compile(r'Fa[\d\/]+')
                fa = fa_re.search(trk)
                if te:
                    int_te = trk.replace('Te','TenGigabitEthernet')
                    list_trk.append(int_te)
                elif ge:
                    int_te = trk.replace('Gi','GigabitEthernet')
                    list_trk.append(trk)
                elif fa:
                    int_te = trk.replace('Fa','FastEthernet')
                    list_trk.append(trk)
       
            mac_cmd = 'show mac address-table dynamic | exclude %s' % '|'.join(list_trk)
            mac_output = self.command(mac_cmd)
            mac_re = re.compile(r'\w+\.\w+\.\w+')
            int_re = re.compile(r'(FastEthernet[\/\d]+|GigabitEthernet[\/\d]+|TengigabitEthernet[\/\d]+)')
            vlan_re = re.compile(r'(\d+|Te[\d\/]+|Gi[\/\d]+|Fa[\/\d+].*)')
            mac_line = mac_output.splitlines()
            mac_int_list = list()

            for line in mac_line:
                data_mac = {
                        'mac' : None,
                        'port' : None,
                        'vlan' : None,
                        'switch' : None,
                        'addr' : None
                        }

                mac = mac_re.search(line)
                port = int_re.search(line)
                vlan = vlan_re.search(line)

                if mac and port and vlan:
                    data_mac['mac'] = mac.group()
                    data_mac['port'] =port.group(1)
                    data_mac['vlan'] = vlan.group(1)
                    data_mac['switch'] = self.hostname

                    mac_int_list.append(data_mac)
            return mac_int_list

        else:
            mac_cmd = 'show mac address-table dynamic | exclude %s' % '|'.join(trk_list)
            mac_output = self.command(mac_cmd)
            mac_re = re.compile(r'\w+\.\w+\.\w+')
            int_re = re.compile(r'(Fa[\/\d]+|Gi[\/\d]+|Te[\/\d]+)')
            vlan_re = re.compile(r'(\d+).*')
            mac_line = mac_output.splitlines()
            mac_int_list = list()
            
            for line in mac_line:
                data_mac = {
                        'mac' : None,
                        'port' : None,
                        'vlan' : None,
                        'switch' : None,
                        'addr': None
                        }
                mac = mac_re.search(line)
                port = int_re.search(line)
                vlan = vlan_re.search(line)

                if mac and port and vlan:
                    data_mac['mac'] = mac.group()
                    data_mac['port'] = port.group(1)
                    data_mac['vlan'] = vlan.group(1)
                    data_mac['switch'] = self.hostname

                    mac_int_list.append(data_mac)

            return mac_int_list

    def trunk_port(self):
        local_cmd = 'show interface trunk | i trunking'
        trunk_cmd = self.command(local_cmd)
        trk_int_re = re.compile(r'(\w+[\/\d]+).*')
        trk_int = trk_int_re.findall(trunk_cmd)

        return trk_int

    def sw_version(self):
        local_cmd = 'show version'
        result = self.command(local_cmd)

        if self.dev_type == 'cisco_ios':
            ios_sw_re = re.compile(r'Version (\S+),')
            ios_pn_re = re.compile(r'Model\s+number\s+:\s+(\S+)')

            ios_sw = ios_sw_re.search(result)
            ios_pn = ios_pn_re.search(result)

            if ios_sw:
                sw = ios_sw.group(1)
                if ios_pn:
                    pn = ios_pn.group(1)
                    return sw,pn
        elif self.dev_type == 'cisco_xe':
            xe_sw_re = re.compile(r'Version (\S+)')
            xe_pn_re = re.compile(r'cisco (WS-\S+)')

            xe_sw = xe_sw_re.search(result)
            xe_pn = xe_pn_re.search(result)

            if xe_sw:
                sw = xe_sw.group(1)
                if xe_pn:
                    pn = xe_pn.group(1)
                    return sw,pn
        elif self.dev_type == 'cisco_nx':
            nx_sw_re = re.compile(r'system:\s+version\s+(\S+)')
            nx_pn_re = re.compile(r'cisco (Nexus\w+ \S+)')

            nx_sw = nx_sw_re.search(result)
            nx_pn = nx_pn_re.search(result)

            if nx_sw:
                sw = nx_sw.group(1)
                if nx_pn:
                    pn = nx_pn.group(1)
                    return sw,pn

def int_vlan(output):
    intvlan_re = re.compile(r'Vlan(\d+)')
    intvlan = intvlan_re.findall(output)

    return intvlan


def main():
    invent_file = 'data/inventory.xlsx'
    data_log = 'syslog/log.txt'
    while True:
        print('\n\n\nthis tools is used to help you to collect data related with network')
        print('this tools is still limited based on Cisco device, Sorry :(')
        print('please make sure you have fill inventory data in "data" directory')
        print('[1] collect software version')
        print('[2] collect address based on arp')
        print('[3] bandiwdth usage')
        print('[4] capture vlan database')
        print('[q] exit\n\n')
        input_select = input('please select function above :')
        input_select = str(input_select)

        if input_select == 'q' or input_select == 'Q':
            sys.exit()

        elif input_select == '1':
            '''create file software inventory'''
            wb = Workbook()
            ws = wb.active
            swinvent_dir = 'data/software_inventory.xlsx'
            ws.cell(row=1, column=1, value='Hostname')
            ws.cell(row=1, column=2, value='Part Number')
            ws.cell(row=1, column=3, value='Software Version')

            wb.save(swinvent_dir)
            inventories = inventory(invent_file)
            cmd = 'show version'
            row = 2
            for d in inventories:
                output = SendCommand(d['hostname'], d['type'], d['address'], d['user'], d['pass'])
                sw_inventory = output.sw_version() 
                print('%s : %s --> partnumber : %s, sw version : %s' % (time_log(), d['hostname'], sw_inventory[1], sw_inventory[0]))
                log('%s : %s --> %s\n' % (time_log(), d['hostname'], sw_inventory), data_log) 
                
                '''load workbook'''
                wb = load_workbook(swinvent_dir)
                ws = wb.active
                ws.cell(row=row, column=1, value=d['hostname'])
                ws.cell(row=row, column=2, value=sw_inventory[1])
                ws.cell(row=row, column=3, value=sw_inventory[0])
                wb.save(swinvent_dir)
                row += 1

        elif input_select == '2':
            dev_inventories = inventory(invent_file)
            address_invent = 'data/address_inventory.xlsx'
            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=1, value='Hostname')
            ws.cell(row=1, column=2, value='Port')
            ws.cell(row=1, column=3, value='Mac Address')
            ws.cell(row=1, column=4, value='Vlan')
            ws.cell(row=1, column=5, value='IP Address')
            wb.save(address_invent)

            int_br_cmd = 'show ip interface brief | include .+\..+\..+\..+'
            inventories = list()
            print('identifying interface vlan, please wait')
            for a in dev_inventories:
                try:
                    int_br_conn = SendCommand(
                            hostname=a['hostname'],
                            dev_type=a['type'],
                            username=a['user'],
                            password=a['pass'],
                            address=a['address']
                            )
                    
                    int_br_output = int_br_conn.command(int_br_cmd)
                    int_vlans = int_vlan(int_br_output)
                    if len(int_vlans) != 0:
                        a['int_vlan'] = int_vlans 
                    
                    inventories.append(a)
                except:
                    continue

            #pprint(inventories)

            for d in inventories:
                output = SendCommand(d['hostname'], d['type'], d['address'], d['user'], d['pass'])
                macs = output.mac_port()
                #pprint('%s -> %s '% (d['hostname'], macs))
                row = 2
                for mac in macs:
                    mac_add = mac['mac']
                    
                    for e in inventories:
                        if mac['vlan'] in e['int_vlan']:
                            arp_cmd = 'show ip arp | include %s' % mac_add
                            #print('connect to %s' % e['hostname'])
                            conn = SendCommand(e['hostname'], e['type'], e['address'], e['user'], e['pass'])
                            arp_output = conn.command(arp_cmd)
                            if arp_output == None:
                                continue
                            else:
                                arp = arp_addr(arp_output)
                                mac['addr'] = arp
                                break
                        else:
                            continue
                    print(mac)
                    wb = load_workbook(address_invent)
                    ws = wb.active
                    ws.cell(row=row, column=1, value=mac['switch'])
                    ws.cell(row=row, column=2, value=mac['port'])
                    ws.cell(row=row, column=3, value=mac['mac'])
                    ws.cell(row=row, column=4, value=mac['vlan'])
                    if mac['addr'] != None:
                        ws.cell(row=row, column=5, value=','.join(mac['addr']))

                    wb.save(address_invent)
                    row += 1
                    log('%s : %s\n' % (time_log(),mac), data_log)

        elif input_select == '3':
            traffic_file = 'data/traffic.xlsx'
            wb = Workbook()
            wb.save(traffic_file)
            device_inventories = inventory(invent_file)
            for host in device_inventories:
                host_conn = SendCommand(host['hostname'], host['type'], host['address'], host['user'], host['pass'])
                print('Collecting traffic rate')
                rates = host_conn.traffic_interface()
                wb = load_workbook(traffic_file)
                ws = wb.create_sheet(host['hostname'])
                ws.cell(row=1, column=1, value='interface')
                ws.cell(row=1, column=2, value='input rate')
                ws.cell(row=1, column=3, value='output rate')
                row = 2
                for rate in rates:
                    ws.cell(row=row, column=1, value=rate['interface'])
                    ws.cell(row=row, column=2, value=int(rate['input']))
                    ws.cell(row=row, column=3, value=int(rate['output']))
                    row += 1
                ws.cell(row=row, column=1, value='Total')
                wb.save(traffic_file)
                print('Collecting data traffic successfully')
        elif input_select == '4':
            vlan_inventory = 'data/vlan_database.xlsx'
            template = open('template/cisco_show_vlan.template')
            result_template = textfsm.TextFSM(template)
            device_inventories = inventory(invent_file)
            vlan_list = list()
            for host in device_inventories:
                try:
                    host_conn = SendCommand(host['hostname'], host['type'], host['address'], host['user'], host['pass'])
                    sh_vlan_cmd = 'show vlan'
                    output = host_conn.command(sh_vlan_cmd)
                    vlan_data = result_template.ParseText(output)
                    for vlan in vlan_data:
                        vlan_dict = {
                                    'id' : None,
                                    'name' : None,
                                    'status' : None
                                }
                        vlan_dict['id'] = vlan[0]
                        vlan_dict['name'] = vlan[1]
                        vlan_dict['status'] = vlan[2]
                        vlan_list.append(vlan_dict)
                    print('%s : %s OK' % (time_log(), host['hostname']))
                    log('%s OK\n' % host['hostname'] ,data_log)
                except:
                    print('%s : %s NOK' % (time_log(), host['hostname']))
                    log('%s NOK\n' % host['hostname'] ,data_log)
            vlan_set = list({d['id'] : d for d in vlan_list}.values())
            print('saving document ..')
            wb = Workbook()
            ws = wb.active

            ws.cell(row=1, column=1, value='Vlan ID')
            ws.cell(row=1, column=2, value='Name')
            ws.cell(row=1, column=3, value='Status')

            row=2
            for data in vlan_set:
                ws.cell(row=row, column=1, value=data['id'])
                ws.cell(row=row, column=2, value=data['name'])
                ws.cell(row=row, column=3, value=data['status'])
                
            wb.save(vlan_inventory)
        else:
            print('sorry, please choose function above..')
            time.sleep(3)

if __name__ == '__main__':
    main()
